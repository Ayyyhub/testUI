import openpyxl

from typing import List

from pages.TestData import AETestData


class Excellreader:
    def __init__(self, file_path: str):
        self.file_path = file_path
        # self.file_path = "D:/AE_PythonProject/testUI/testcases/test_data.xlsx" #代码失去灵活性，变成 “专用工具”
        self.workbook = openpyxl.load_workbook(self.file_path)

    """从Excel读取测试数据"""
    def get_test_data(self, sheet_name: str) -> List[AETestData]:

        # 打印所有可用的工作表名称
        # print("实际读取到的工作表：", self.workbook.sheetnames)
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(
                f"工作表 '{sheet_name}' 不存在，可用工作表：{self.workbook.sheetnames}"
            )
        sheet = self.workbook[sheet_name]
        test_data_list = []

        # 跳过标题行，从第二行开始读取
        for row in range(2, sheet.max_row + 1):
            # 获取测试用例ID并检查是否被注释
            test_case_id = sheet.cell(row=row, column=1).value
            if test_case_id and not str(test_case_id).strip().startswith("//"):
                test_data = AETestData(
                    test_case_id=test_case_id,
                    description=sheet.cell(row=row, column=2).value,
                    step_id=sheet.cell(row=row, column=3).value,
                    determin_type=sheet.cell(row=row, column=4).value,
                    determin_method=sheet.cell(row=row, column=5).value,
                    determin_value=sheet.cell(row=row, column=6).value,
                    input_value=sheet.cell(row=row, column=7).value,
                    cv_points=sheet.cell(row=row, column=8).value,
                    assert_type=sheet.cell(row=row, column=9).value,
                    assert_method=sheet.cell(row=row, column=10).value,
                    expected_result=sheet.cell(row=row, column=11).value,
                    outputed_result=sheet.cell(row=row, column=12).value,
                    status=sheet.cell(row=row, column=13).value,
                )
                test_data_list.append(test_data)

        return test_data_list

    # def write_result(self, test_case_id: str, actual_result: str, status: str):
    #     """将测试结果写回Excel"""
    #     sheet = self.workbook.active
    #     for row in range(2, sheet.max_row + 1):
    #         if sheet.cell(row=row, column=1).value == test_case_id:
    #             sheet.cell(row=row, column=6).value = actual_result  # 实际结果
    #             sheet.cell(row=row, column=7).value = status  # 测试状态
    #             break
    #     self.workbook.save(self.file_path)
